#!/usr/bin/env python3
"""
amazon_research.py - Amazon Market Research Analysis Script

Modes:
  merge           - Merge raw Smacient search JSON files, deduplicate by ASIN
  filter          - Use Gemini to classify products as RELEVANT / BORDERLINE / IRRELEVANT
  apply-decisions - Apply user approve/reject decisions for borderline products
  enrich          - Use Gemini to extract brand names and product formats
  analyze         - Full market analysis -> Excel + Markdown summary

Usage:
  python scripts/amazon_research.py --mode merge --workdir path/to/workdir
  python scripts/amazon_research.py --mode filter --category "X" --workdir path
  python scripts/amazon_research.py --mode apply-decisions --workdir path
  python scripts/amazon_research.py --mode enrich --category "X" --workdir path
  python scripts/amazon_research.py --mode analyze --category "X" --marketplace in --workdir path
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path


def load_env():
    env_path = Path(__file__).parent.parent / ".env"
    if env_path.exists():
        with open(env_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    os.environ.setdefault(key.strip(), val.strip())


load_env()


# ---------------------------------------------------------------------------
# MODE: MERGE
# ---------------------------------------------------------------------------

def mode_merge(workdir):
    import pandas as pd

    workdir = Path(workdir)
    raw_files = sorted(workdir.glob("raw_search_*.json"))

    if not raw_files:
        print("ERROR: No raw_search_*.json files found in workdir")
        sys.exit(1)

    all_records = []
    query_map = {}  # asin -> list of queries

    for fpath in raw_files:
        with open(fpath, encoding="utf-8") as f:
            data = json.load(f)

        # Handle both direct list and wrapped {"results": [...]} formats
        if isinstance(data, dict):
            products = data.get("results", data.get("data", data.get("products", [])))
        else:
            products = data

        for product in products:
            asin = str(product.get("asin", "")).strip()
            if not asin:
                continue

            if asin not in query_map:
                query_map[asin] = []
            q = str(product.get("query", product.get("search_term", ""))).strip()
            if q and q not in query_map[asin]:
                query_map[asin].append(q)

            all_records.append(product)

    if not all_records:
        print("ERROR: No products found across all raw search files")
        sys.exit(1)

    df = pd.DataFrame(all_records)

    # Normalize asin column name
    asin_col = next((c for c in df.columns if c.lower() == "asin"), None)
    if asin_col and asin_col != "asin":
        df = df.rename(columns={asin_col: "asin"})

    if "asin" not in df.columns:
        print("ERROR: No 'asin' column found in data")
        print(f"Available columns: {list(df.columns)}")
        sys.exit(1)

    df["asin"] = df["asin"].astype(str).str.strip()
    df = df[df["asin"] != ""].drop_duplicates(subset="asin", keep="first")
    df["all_queries"] = df["asin"].map(lambda a: " | ".join(query_map.get(a, [])))

    out_path = workdir / "merged.csv"
    df.to_csv(out_path, index=False)

    print(f"MERGE COMPLETE")
    print(f"Total records across all searches: {len(all_records)}")
    print(f"Unique ASINs after dedup: {len(df)}")
    print(f"Columns: {list(df.columns)}")
    print(f"Output: {out_path}")


# ---------------------------------------------------------------------------
# MODE: FILTER
# ---------------------------------------------------------------------------

def mode_filter(workdir, category):
    from google import genai
    import pandas as pd

    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: GEMINI_API_KEY not set. Check .env file.")
        sys.exit(1)

    client = genai.Client(api_key=api_key)
    GEMINI_MODEL = "gemini-2.5-flash"

    workdir = Path(workdir)
    df = pd.read_csv(workdir / "merged.csv")
    titles = df["title"].fillna("").tolist()
    asins = df["asin"].tolist()

    classifications = []
    BATCH_SIZE = 25

    print(f"Filtering {len(titles)} products for category: '{category}'")

    for i in range(0, len(titles), BATCH_SIZE):
        batch_titles = titles[i : i + BATCH_SIZE]
        batch_asins = asins[i : i + BATCH_SIZE]
        numbered = "\n".join(f"{j+1}. {t}" for j, t in enumerate(batch_titles))

        prompt = f"""You are helping with Amazon market research. Classify each product listing for the category: "{category}"

For each product, return exactly one of:
- RELEVANT: Clearly fits the category (right product type AND right target audience)
- BORDERLINE: Uncertain - could fit but missing key info, or partially matches
- IRRELEVANT: Clearly does not belong (wrong product, wrong audience, or off-category)

For BORDERLINE and IRRELEVANT, add a brief reason (max 8 words).

Return ONLY a JSON array, no other text:
[{{"index": 1, "classification": "RELEVANT"}}, {{"index": 2, "classification": "IRRELEVANT", "reason": "adult product not for kids"}}]

Products to classify:
{numbered}"""

        try:
            response = client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
            raw = response.text.strip()
            if raw.startswith("```"):
                raw = re.sub(r"```(?:json)?", "", raw).strip().rstrip("`").strip()
            batch_results = json.loads(raw)
        except Exception as e:
            print(f"  WARNING: Gemini error for batch {i//BATCH_SIZE + 1}: {e}. Marking all as BORDERLINE.")
            batch_results = [{"index": j+1, "classification": "BORDERLINE", "reason": "Gemini error"} for j in range(len(batch_titles))]

        for r in batch_results:
            idx_in_batch = r.get("index", 0) - 1
            if 0 <= idx_in_batch < len(batch_titles):
                classifications.append({
                    "asin": batch_asins[idx_in_batch],
                    "title": batch_titles[idx_in_batch],
                    "classification": r.get("classification", "BORDERLINE"),
                    "reason": r.get("reason", ""),
                })

        print(f"  Classified {min(i + BATCH_SIZE, len(titles))}/{len(titles)}...")

    relevant = [c for c in classifications if c["classification"] == "RELEVANT"]
    borderline = [c for c in classifications if c["classification"] == "BORDERLINE"]
    irrelevant = [c for c in classifications if c["classification"] == "IRRELEVANT"]

    relevant_asins = {c["asin"] for c in relevant}
    borderline_asins = {c["asin"] for c in borderline}

    df[df["asin"].isin(relevant_asins)].to_csv(workdir / "relevant.csv", index=False)
    df[df["asin"].isin(borderline_asins)].to_csv(workdir / "borderline_data.csv", index=False)

    with open(workdir / "borderline.json", "w", encoding="utf-8") as f:
        json.dump(borderline, f, indent=2, ensure_ascii=False)

    print(f"\nFILTER COMPLETE")
    print(f"RELEVANT: {len(relevant)}")
    print(f"BORDERLINE: {len(borderline)}")
    print(f"IRRELEVANT: {len(irrelevant)}")

    if irrelevant:
        samples = [c["title"][:70] for c in irrelevant[:5]]
        print(f"Sample irrelevant: {samples}")


# ---------------------------------------------------------------------------
# MODE: APPLY-DECISIONS
# ---------------------------------------------------------------------------

def mode_apply_decisions(workdir):
    import pandas as pd

    workdir = Path(workdir)

    relevant_df = pd.read_csv(workdir / "relevant.csv")

    decisions_path = workdir / "user_decisions.json"
    if not decisions_path.exists():
        # No borderline products - just copy relevant to approved
        relevant_df.to_csv(workdir / "approved.csv", index=False)
        print(f"APPLY_DECISIONS COMPLETE")
        print(f"No borderline decisions found. Final product count: {len(relevant_df)}")
        return

    with open(decisions_path, encoding="utf-8") as f:
        decisions = json.load(f)

    approved_asins = {d["asin"] for d in decisions if d.get("approved")}

    borderline_data_path = workdir / "borderline_data.csv"
    if borderline_data_path.exists() and approved_asins:
        borderline_df = pd.read_csv(borderline_data_path)
        approved_borderline = borderline_df[borderline_df["asin"].isin(approved_asins)]
        final_df = pd.concat([relevant_df, approved_borderline], ignore_index=True)
    else:
        final_df = relevant_df

    final_df.to_csv(workdir / "approved.csv", index=False)

    print(f"APPLY_DECISIONS COMPLETE")
    print(f"Base relevant: {len(relevant_df)}")
    print(f"Borderlines approved: {len(final_df) - len(relevant_df)}")
    print(f"Final product count: {len(final_df)}")


# ---------------------------------------------------------------------------
# MODE: ENRICH (Brand + Format via Gemini)
# ---------------------------------------------------------------------------

def mode_enrich(workdir, category):
    from google import genai
    import pandas as pd

    api_key = os.environ.get("GEMINI_API_KEY")
    if not api_key:
        print("ERROR: GEMINI_API_KEY not set. Check .env file.")
        sys.exit(1)

    client = genai.Client(api_key=api_key)
    GEMINI_MODEL = "gemini-2.5-flash"

    workdir = Path(workdir)
    df = pd.read_csv(workdir / "approved.csv")
    df = df.reset_index(drop=True)
    titles = df["title"].fillna("").tolist()

    # Step 1: Detect relevant format categories for this product type
    format_prompt = f"""For Amazon products in the category "{category}", what are the 4-6 most common product format/form types?

Examples:
- "protein powder for women" -> ["Powder", "Bar", "Ready-to-drink", "Capsules", "Other"]
- "melatonin for kids" -> ["Gummies", "Syrup", "Drops", "Tablets", "Spray", "Other"]
- "baby shampoo" -> ["Shampoo", "2-in-1 Shampoo+Conditioner", "Foam", "Bar", "Other"]

Return ONLY a JSON array of 4-6 format strings. Always include "Other" as the last option."""

    try:
        fmt_response = client.models.generate_content(model=GEMINI_MODEL, contents=format_prompt)
        raw = fmt_response.text.strip()
        if raw.startswith("```"):
            raw = re.sub(r"```(?:json)?", "", raw).strip().rstrip("`").strip()
        detected_formats = json.loads(raw)
    except Exception:
        detected_formats = ["Tablets", "Capsules", "Gummies", "Powder", "Liquid", "Other"]

    formats_str = ", ".join(detected_formats)
    print(f"Detected formats for '{category}': {formats_str}")

    # Step 2: Enrich each product with brand + format
    brand_map = {}
    format_map = {}
    BATCH_SIZE = 10
    MAX_RETRIES = 3

    print(f"Enriching {len(titles)} products...")

    for i in range(0, len(titles), BATCH_SIZE):
        batch = titles[i : i + BATCH_SIZE]
        numbered = "\n".join(f"{j+1}. {t}" for j, t in enumerate(batch))

        prompt = f"""For each Amazon product title, extract two things:

1. brand_name: The brand as printed on the product packaging (not the product description or ingredients).
   - Look for the brand at the START or END of the title
   - Do NOT use generic descriptors like "Natural", "Organic", "Kids" as brands
   - If truly no brand is identifiable, use "Unbranded"

2. format: The product's physical form. Choose from: {formats_str}

Category context: "{category}"

Return ONLY a valid JSON array with exactly {len(batch)} objects, one per product:
[{{"index": 1, "brand_name": "BrandX", "format": "Liquid Wash"}}, {{"index": 2, "brand_name": "BrandY", "format": "Other"}}]

IMPORTANT: Return valid JSON only. No extra text, no markdown, no trailing commas.

Products:
{numbered}"""

        batch_results = None
        last_err = None
        for attempt in range(MAX_RETRIES):
            try:
                response = client.models.generate_content(model=GEMINI_MODEL, contents=prompt)
                raw = response.text.strip()
                if raw.startswith("```"):
                    raw = re.sub(r"```(?:json)?", "", raw).strip().rstrip("`").strip()
                batch_results = json.loads(raw)
                break
            except Exception as e:
                last_err = e
                if attempt < MAX_RETRIES - 1:
                    print(f"  Retry {attempt + 1}/{MAX_RETRIES - 1} for batch {i//BATCH_SIZE + 1}: {e}")

        if batch_results is None:
            print(f"  ERROR: Batch {i//BATCH_SIZE + 1} failed after {MAX_RETRIES} attempts: {last_err}")
            print(f"  Titles in failed batch:")
            for j, t in enumerate(batch):
                print(f"    {i+j+1}. {t[:80]}")
            batch_results = [{"index": j+1, "brand_name": "PARSE_FAILED", "format": "Other"} for j in range(len(batch))]

        for r in batch_results:
            idx_in_batch = r.get("index", 0) - 1
            global_idx = i + idx_in_batch
            if 0 <= global_idx < len(df):
                asin = df.iloc[global_idx]["asin"]
                brand_map[str(global_idx)] = r.get("brand_name", "Unbranded")
                format_map[str(asin)] = r.get("format", "Other")

        print(f"  Enriched {min(i + BATCH_SIZE, len(titles))}/{len(titles)}...")

    with open(workdir / "brand_mapping.json", "w", encoding="utf-8") as f:
        json.dump(brand_map, f, indent=2, ensure_ascii=False)

    with open(workdir / "format_mapping.json", "w", encoding="utf-8") as f:
        json.dump(format_map, f, indent=2, ensure_ascii=False)

    with open(workdir / "formats.json", "w", encoding="utf-8") as f:
        json.dump(detected_formats, f, ensure_ascii=False)

    unique_brands = len(set(brand_map.values()))
    print(f"\nENRICH COMPLETE")
    print(f"Products enriched: {len(brand_map)}")
    print(f"Unique brands found: {unique_brands}")
    print(f"Formats: {formats_str}")


# ---------------------------------------------------------------------------
# MODE: ANALYZE
# ---------------------------------------------------------------------------

def mode_analyze(workdir, category, marketplace):
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    workdir = Path(workdir)

    df = pd.read_csv(workdir / "approved.csv")
    df = df.reset_index(drop=True)

    with open(workdir / "brand_mapping.json", encoding="utf-8") as f:
        brand_map = json.load(f)

    with open(workdir / "format_mapping.json", encoding="utf-8") as f:
        format_map = json.load(f)

    with open(workdir / "formats.json", encoding="utf-8") as f:
        detected_formats = json.load(f)

    # --- Normalize column names from Smacient's output ---
    rename = {}
    for col in df.columns:
        lc = col.lower().replace(" ", "_").replace("-", "_")
        if col in ("asin", "title", "query", "all_queries"):
            continue
        elif lc in ("price", "selling_price", "current_price", "sale_price") and "list" not in lc:
            rename[col] = "price"
        elif lc in ("list_price", "mrp", "original_price", "was_price", "rrp"):
            rename[col] = "list_price"
        elif lc in ("rating", "star_rating", "avg_rating", "stars"):
            rename[col] = "rating"
        elif lc in ("review_count", "reviews", "num_reviews", "number_of_reviews", "ratings_total"):
            rename[col] = "review_count"
        elif any(x in lc for x in ("bought_past", "bought_in_past", "monthly_sales", "sales_volume", "units_sold", "monthly_units", "sold_last")):
            rename[col] = "bought_past_month"
        elif any(x in lc for x in ("is_sponsored", "sponsored")):
            rename[col] = "is_sponsored"
        elif any(x in lc for x in ("discount", "discount_pct", "savings_pct", "off_pct")):
            rename[col] = "discount_pct"
        elif lc in ("is_prime", "prime"):
            rename[col] = "is_prime"

    df = df.rename(columns=rename)

    # Ensure required columns exist with safe defaults
    for col in ["price", "list_price", "rating", "review_count", "bought_past_month", "is_sponsored", "discount_pct"]:
        if col not in df.columns:
            df[col] = None

    for col in ["price", "list_price", "discount_pct", "rating", "review_count"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # Assign brand and format
    df["brand"] = [brand_map.get(str(i), "Unknown") for i in df.index]
    df["format"] = df["asin"].map(lambda a: format_map.get(str(a), "Other"))

    # --- Units midpoint conversion ---
    MIDPOINTS = {
        "50+": 75, "100+": 150, "200+": 250, "300+": 350, "400+": 450,
        "500+": 600, "600+": 700, "700+": 800, "800+": 900, "900+": 1000,
        "1k+": 1500, "1K+": 1500, "2k+": 2500, "2K+": 2500, "3k+": 3500, "3K+": 3500,
        "4k+": 5000, "4K+": 5000, "5k+": 6250, "5K+": 6250, "6k+": 7500, "6K+": 7500,
        "7k+": 8500, "7K+": 8500, "8k+": 9000, "8K+": 9000, "10k+": 12000, "10K+": 12000,
        "20k+": 25000, "20K+": 25000, "50k+": 60000, "50K+": 60000,
    }

    def to_midpoint(val):
        if val is None or (isinstance(val, float) and str(val) == "nan"):
            return None
        val_str = str(val).strip()
        if val_str in MIDPOINTS:
            return MIDPOINTS[val_str]
        # Some tools return plain integers
        try:
            return float(val_str)
        except (ValueError, TypeError):
            return None

    df["units_midpoint"] = df["bought_past_month"].apply(to_midpoint)
    df["est_monthly_revenue"] = df.apply(
        lambda r: r["units_midpoint"] * r["price"]
        if (r["units_midpoint"] is not None and not (isinstance(r["units_midpoint"], float) and str(r["units_midpoint"]) == "nan")
            and r["price"] is not None and not (isinstance(r["price"], float) and str(r["price"]) == "nan"))
        else None,
        axis=1,
    )

    # --- Price tiers ---
    def price_tier(p):
        if p is None or (isinstance(p, float) and str(p) == "nan"):
            return "Unknown"
        if p < 300: return "Budget (<300)"
        if p < 600: return "Mid (300-600)"
        if p < 1000: return "Premium (601-1000)"
        return "Super-premium (>1000)"

    df["price_tier"] = df["price"].apply(price_tier)

    # --- Summary metrics ---
    total_products = len(df)
    products_with_sales = df["units_midpoint"].notna().sum()
    sales_coverage_pct = (products_with_sales / total_products * 100) if total_products > 0 else 0
    visible_revenue = df["est_monthly_revenue"].sum() or 0
    extrapolated_revenue = (visible_revenue / (sales_coverage_pct / 100)) if sales_coverage_pct > 0 else 0
    total_brands = df["brand"].nunique()
    median_price = df["price"].median()
    mean_price = df["price"].mean()
    avg_rating = df["rating"].mean()

    def is_sponsored(x):
        return str(x).lower().strip() in ("true", "1", "yes", "y")

    sponsored_count = df["is_sponsored"].apply(is_sponsored).sum()
    sponsored_pct = (sponsored_count / total_products * 100) if total_products > 0 else 0

    def fmt_inr(val):
        if val is None or val == 0:
            return "N/A"
        if val >= 1e7:
            return f"INR {val/1e7:.2f} Cr"
        elif val >= 1e5:
            return f"INR {val/1e5:.2f} L"
        else:
            return f"INR {val:,.0f}"

    # --- Brand analysis ---
    brand_grp = df.groupby("brand").agg(
        product_count=("asin", "count"),
        avg_price=("price", "mean"),
        min_price=("price", "min"),
        max_price=("price", "max"),
        avg_rating=("rating", "mean"),
        avg_discount=("discount_pct", "mean"),
        est_monthly_units=("units_midpoint", "sum"),
        est_monthly_revenue=("est_monthly_revenue", "sum"),
    ).reset_index()
    brand_grp["market_share_pct"] = (
        brand_grp["est_monthly_revenue"] / visible_revenue * 100
        if visible_revenue > 0 else 0
    )
    brand_grp = brand_grp.sort_values("est_monthly_revenue", ascending=False)

    # --- Top products ---
    rev_cols = ["brand", "title", "price", "units_midpoint", "est_monthly_revenue", "rating", "review_count", "format", "asin"]
    top_revenue = (
        df[df["est_monthly_revenue"].notna()]
        .nlargest(20, "est_monthly_revenue")[[c for c in rev_cols if c in df.columns]]
    )

    units_cols = ["brand", "title", "units_midpoint", "price", "est_monthly_revenue", "rating", "format", "asin"]
    top_units = (
        df[df["units_midpoint"].notna()]
        .nlargest(20, "units_midpoint")[[c for c in units_cols if c in df.columns]]
    )

    rating_cols = ["brand", "title", "rating", "review_count", "price", "units_midpoint", "format", "asin"]
    has_reviews = df["review_count"].notna() & (df["review_count"] >= 10)
    if has_reviews.sum() >= 5:
        top_rating = df[has_reviews].nlargest(20, "rating")[[c for c in rating_cols if c in df.columns]]
    else:
        top_rating = df.nlargest(20, "rating")[[c for c in rating_cols if c in df.columns]]

    discount_cols = ["brand", "title", "price", "list_price", "discount_pct", "rating", "format", "asin"]
    top_discount = (
        df[df["discount_pct"].notna()]
        .nlargest(20, "discount_pct")[[c for c in discount_cols if c in df.columns]]
    )

    # --- Format breakdown ---
    fmt_grp = df.groupby("format").agg(
        product_count=("asin", "count"),
        avg_price=("price", "mean"),
        avg_rating=("rating", "mean"),
        est_monthly_revenue=("est_monthly_revenue", "sum"),
    ).reset_index()
    fmt_grp["revenue_share_pct"] = (
        fmt_grp["est_monthly_revenue"] / visible_revenue * 100
        if visible_revenue > 0 else 0
    )
    fmt_grp = fmt_grp.sort_values("product_count", ascending=False)

    # --- Pricing analysis ---
    tier_grp = df.groupby("price_tier").agg(
        product_count=("asin", "count"),
        avg_rating=("rating", "mean"),
        avg_discount=("discount_pct", "mean"),
        est_monthly_units=("units_midpoint", "sum"),
        est_monthly_revenue=("est_monthly_revenue", "sum"),
    ).reset_index()
    tier_grp["revenue_share_pct"] = (
        tier_grp["est_monthly_revenue"] / visible_revenue * 100
        if visible_revenue > 0 else 0
    )
    tier_order = ["Budget (<300)", "Mid (300-600)", "Premium (601-1000)", "Super-premium (>1000)", "Unknown"]
    tier_grp["_order"] = tier_grp["price_tier"].map({t: i for i, t in enumerate(tier_order)}).fillna(99)
    tier_grp = tier_grp.sort_values("_order").drop(columns=["_order"])

    # --- Competitive quadrants ---
    units_q75 = df["units_midpoint"].quantile(0.75)
    rating_q75 = df["rating"].quantile(0.75)

    def quadrant(row):
        high_u = (row["units_midpoint"] is not None
                  and not (isinstance(row["units_midpoint"], float) and str(row["units_midpoint"]) == "nan")
                  and row["units_midpoint"] >= units_q75)
        high_r = (row["rating"] is not None
                  and not (isinstance(row["rating"], float) and str(row["rating"]) == "nan")
                  and row["rating"] >= rating_q75)
        if high_u and high_r: return "Established Winner"
        if high_u: return "Volume Player (Quality Gap)"
        if high_r: return "High Quality / Low Visibility"
        return "Emerging / Niche"

    df["competitive_quadrant"] = df.apply(quadrant, axis=1)

    flags_cols = ["brand", "title", "price", "units_midpoint", "rating", "review_count",
                  "est_monthly_revenue", "competitive_quadrant", "discount_pct", "asin"]
    flags_df = df[[c for c in flags_cols if c in df.columns]].sort_values("competitive_quadrant")

    # --- Excel output ---
    output_dir = Path(__file__).parent.parent / "outputs"
    output_dir.mkdir(exist_ok=True)

    safe_category = re.sub(r'[\\/*?:"<>|]', "", category).strip()
    excel_path = output_dir / f"Amazon - {safe_category.title()} - Market Analysis.xlsx"
    summary_path = output_dir / f"Amazon - {safe_category.title()} - Summary.md"

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    ALT_FILL = PatternFill("solid", fgColor="EEF2FF")

    def write_df_to_sheet(ws, data_df, rename_cols=None):
        display_df = data_df.copy()
        if rename_cols:
            display_df.columns = rename_cols

        headers = list(display_df.columns)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        for r_idx, row in enumerate(display_df.itertuples(index=False), start=2):
            row_vals = []
            for v in row:
                if isinstance(v, float):
                    if str(v) == "nan": row_vals.append(None)
                    else: row_vals.append(round(v, 2))
                else:
                    row_vals.append(v)
            ws.append(row_vals)
            if r_idx % 2 == 0:
                for cell in ws[r_idx]:
                    cell.fill = ALT_FILL

        for col_idx in range(1, len(headers) + 1):
            max_len = max(
                (len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    wb = Workbook()
    wb.remove(wb.active)

    # Tab 1: Market Overview
    ws1 = wb.create_sheet("Market Overview")
    overview = [
        ("Category", safe_category),
        ("Marketplace", f"amazon.{marketplace}"),
        ("Analysis Date", datetime.now().strftime("%Y-%m-%d")),
        ("", ""),
        ("--- PRODUCTS ---", ""),
        ("Total unique products", total_products),
        ("Total unique brands", total_brands),
        ("Products with sales data", f"{products_with_sales} ({sales_coverage_pct:.1f}%)"),
        ("Sponsored products", f"{int(sponsored_count)} ({sponsored_pct:.1f}%)"),
        ("", ""),
        ("--- PRICING ---", ""),
        ("Median selling price", f"INR {median_price:,.0f}" if not (isinstance(median_price, float) and str(median_price) == "nan") else "N/A"),
        ("Mean selling price", f"INR {mean_price:,.0f}" if not (isinstance(mean_price, float) and str(mean_price) == "nan") else "N/A"),
        ("Price range", f"INR {df['price'].min():,.0f} - INR {df['price'].max():,.0f}"),
        ("", ""),
        ("--- MARKET SIZE (MONTHLY EST.) ---", ""),
        ("Visible revenue (products with sales data)", fmt_inr(visible_revenue)),
        ("Extrapolated total market", fmt_inr(extrapolated_revenue)),
        ("Sales data coverage", f"{sales_coverage_pct:.1f}% of products have sales data"),
        ("Extrapolation note", "Assumes non-reporting products sell at similar rates"),
        ("", ""),
        ("--- QUALITY ---", ""),
        ("Average product rating", f"{avg_rating:.2f} / 5" if not (isinstance(avg_rating, float) and str(avg_rating) == "nan") else "N/A"),
    ]

    ws1.cell(row=1, column=1, value="Metric").font = HEADER_FONT
    ws1.cell(row=1, column=1).fill = HEADER_FILL
    ws1.cell(row=1, column=2, value="Value").font = HEADER_FONT
    ws1.cell(row=1, column=2).fill = HEADER_FILL

    for r_idx, (k, v) in enumerate(overview, start=2):
        ws1.cell(row=r_idx, column=1, value=k)
        ws1.cell(row=r_idx, column=2, value=v)
        if str(k).startswith("---"):
            ws1.cell(row=r_idx, column=1).font = Font(bold=True, color="1F3864")

    ws1.column_dimensions["A"].width = 42
    ws1.column_dimensions["B"].width = 32

    # Tab 2: Brand Analysis
    ws2 = wb.create_sheet("Brand Analysis")
    write_df_to_sheet(ws2, brand_grp, rename_cols=[
        "Brand", "Products", "Avg Price (INR)", "Min Price", "Max Price",
        "Avg Rating", "Avg Discount %", "Est. Monthly Units", "Est. Monthly Revenue (INR)", "Market Share %"
    ])

    # Tab 3: Revenue Estimates
    ws3 = wb.create_sheet("Revenue Estimates")
    rev_est_cols = ["brand", "title", "price", "bought_past_month", "units_midpoint",
                    "est_monthly_revenue", "rating", "format", "price_tier", "asin"]
    rev_df = df[[c for c in rev_est_cols if c in df.columns]].sort_values(
        "est_monthly_revenue", ascending=False, na_position="last"
    )
    write_df_to_sheet(ws3, rev_df, rename_cols=[
        "Brand", "Title", "Price (INR)", "Bought Past Month", "Units Midpoint",
        "Est. Revenue (INR)", "Rating", "Format", "Price Tier", "ASIN"
    ][:len(rev_df.columns)])

    # Tab 4: Top Products (4 side-by-side tables)
    ws4 = wb.create_sheet("Top Products")
    sections = [
        ("Top 20 by Est. Monthly Revenue", top_revenue),
        ("Top 20 by Units Sold", top_units),
        ("Top 20 by Rating", top_rating),
        ("Top 20 by Discount %", top_discount),
    ]
    current_col = 1
    for sec_title, sec_df in sections:
        ws4.cell(row=1, column=current_col, value=sec_title).font = Font(bold=True, color="1F3864", size=12)
        for ci, col_name in enumerate(sec_df.columns, start=current_col):
            c = ws4.cell(row=2, column=ci, value=col_name)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        for ri, row in enumerate(sec_df.itertuples(index=False), start=3):
            for ci, val in enumerate(row, start=current_col):
                ws4.cell(row=ri, column=ci, value=round(val, 2) if isinstance(val, float) and str(val) != "nan" else (None if isinstance(val, float) else val))
        current_col += len(sec_df.columns) + 2

    # Tab 5: Format Breakdown
    ws5 = wb.create_sheet("Format Breakdown")
    write_df_to_sheet(ws5, fmt_grp, rename_cols=[
        "Format", "Product Count", "Avg Price (INR)", "Avg Rating",
        "Est. Monthly Revenue (INR)", "Revenue Share %"
    ])

    # Tab 6: Pricing Analysis
    ws6 = wb.create_sheet("Pricing Analysis")
    write_df_to_sheet(ws6, tier_grp, rename_cols=[
        "Price Tier", "Product Count", "Avg Rating", "Avg Discount %",
        "Est. Monthly Units", "Est. Monthly Revenue (INR)", "Revenue Share %"
    ])

    # Tab 7: Competitive Flags
    ws7 = wb.create_sheet("Competitive Flags")
    write_df_to_sheet(ws7, flags_df)

    # Tab 8: Raw Data
    ws8 = wb.create_sheet("Raw Data")
    raw_out = df.drop(columns=["competitive_quadrant"], errors="ignore")
    write_df_to_sheet(ws8, raw_out)

    wb.save(excel_path)

    # --- Markdown Summary ---
    top_brands_rows = list(brand_grp.head(5).itertuples(index=False))
    top_format = fmt_grp.iloc[0]["format"] if len(fmt_grp) > 0 else "N/A"
    top_tier = (
        tier_grp.sort_values("est_monthly_revenue", ascending=False).iloc[0]["price_tier"]
        if len(tier_grp) > 0 else "N/A"
    )
    winners = df[df["competitive_quadrant"] == "Established Winner"]["brand"].unique().tolist()[:5]

    brand_table = "\n".join(
        f"{i+1}. **{r.brand}** - {fmt_inr(r.est_monthly_revenue)} est. monthly revenue ({r.product_count} products, avg INR {r.avg_price:,.0f})"
        for i, r in enumerate(top_brands_rows)
        if hasattr(r, "brand")
    )

    summary_md = f"""# Amazon Market Research: {safe_category.title()}

**Marketplace:** amazon.{marketplace}
**Generated:** {datetime.now().strftime("%Y-%m-%d")}

---

## Market Snapshot

| Metric | Value |
|--------|-------|
| Unique Products Analyzed | {total_products} |
| Unique Brands | {total_brands} |
| Products with Sales Data | {products_with_sales} ({sales_coverage_pct:.1f}%) |
| Visible Monthly Market | {fmt_inr(visible_revenue)} |
| Extrapolated Total Market | {fmt_inr(extrapolated_revenue)} |
| Median Price | INR {median_price:,.0f} |
| Average Rating | {avg_rating:.2f} / 5 |
| Sponsored Share | {sponsored_pct:.1f}% of listings |

---

## Top Brands by Estimated Revenue

{brand_table}

---

## Key Findings

- **Dominant format:** {top_format} leads by product count
- **Top price tier by revenue:** {top_tier}
- **Established Winners** (high volume + high rating): {', '.join(winners) if winners else 'None identified with sufficient data'}

---

## Caveats

- Sales data (bought_past_month) is only visible for {sales_coverage_pct:.1f}% of products
- Revenue estimates use midpoint values (e.g., "1K+" -> 1,500 units)
- Market size extrapolation assumes untracked products sell at similar rates
- This is a point-in-time snapshot - rankings shift over time

---

*Full 8-tab report: `{excel_path.name}`*
"""

    with open(summary_path, "w", encoding="utf-8") as f:
        f.write(summary_md)

    print(f"\nANALYSIS COMPLETE")
    print(f"Products analyzed: {total_products}")
    print(f"Brands: {total_brands}")
    print(f"Visible monthly market: {fmt_inr(visible_revenue)}")
    print(f"Extrapolated market: {fmt_inr(extrapolated_revenue)}")
    print(f"EXCEL_PATH: {excel_path}")
    print(f"SUMMARY_MD_PATH: {summary_path}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Amazon Market Research Analysis")
    parser.add_argument("--mode", required=True,
                        choices=["merge", "filter", "apply-decisions", "enrich", "analyze"])
    parser.add_argument("--workdir", required=True, help="Working directory for this research session")
    parser.add_argument("--category", default="", help="Product category (required for filter/enrich/analyze)")
    parser.add_argument("--marketplace", default="com", help="Amazon marketplace domain")
    args = parser.parse_args()

    if args.mode == "merge":
        mode_merge(args.workdir)
    elif args.mode == "filter":
        if not args.category:
            print("ERROR: --category is required for filter mode")
            sys.exit(1)
        mode_filter(args.workdir, args.category)
    elif args.mode == "apply-decisions":
        mode_apply_decisions(args.workdir)
    elif args.mode == "enrich":
        if not args.category:
            print("ERROR: --category is required for enrich mode")
            sys.exit(1)
        mode_enrich(args.workdir, args.category)
    elif args.mode == "analyze":
        if not args.category:
            print("ERROR: --category is required for analyze mode")
            sys.exit(1)
        mode_analyze(args.workdir, args.category, args.marketplace)
