"""
Build the Citable Listicle Architecture workbook + markdown report from
the JSON produced by audit_listicles.py.

Usage:
    python build_outputs.py <extract_json> [output_prefix]

Writes:
    <output_prefix>_workbook.xlsx  (Full Data / Summary / Point Detail tabs)
    <output_prefix>_report.md

Requires: openpyxl
"""

import json
import statistics
import sys
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="100739", end_color="100739", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True, size=12)


def compute_scores(r):
    """Each point scores 0-1. Sub-checks that could not be detected
    (None) are excluded from that point's average rather than counted
    as a fail, and are reported separately as a coverage caveat."""

    def avg(vals):
        vals = [v for v in vals if v is not None]
        return sum(vals) / len(vals) if vals else None

    p1 = 1.0 if r["p1_buyer_query_match"] else (0.5 if r["p1_has_best_or_top"] else 0.0)

    p2_subs = [r["p2_pick_count_in_range_5_8"], r["p2_has_best_for_labels"], r["p2_quick_answer_near_top"]]
    p2 = avg(p2_subs)

    p3 = avg(
        [
            r["p3_methodology_heading_present"] or r["p3_criteria_language_present"],
            r["p3_visible_byline_present"],
            r["p3_visible_date_present"],
        ]
    )

    p4 = 1.0 if (r["p4_has_comparison_table"] and r["p4_table_column_match_count_of_8"] >= 3) else (
        0.5 if r["p4_has_comparison_table"] else 0.0
    )

    if r["p5_vendor_sections_present"]:
        p5_pct = avg(
            [
                r["p5_pct_sections_with_features"],
                r["p5_pct_sections_with_limits"],
                r["p5_pct_sections_with_pricing"],
                r["p5_pct_sections_with_choose_if"],
            ]
        )
        p5 = min(p5_pct / 100, 1.0) if p5_pct is not None else None
    else:
        p5 = 0.0 if r["p2_pick_detection_method"] != "not_detected" else None

    p6 = 1.0 if r["p6_decision_framework_present"] else 0.0
    p7 = r["p7_proof_score_of_6"] / 6

    components = [p1, p2, p3, p4, p5, p6, p7]
    scored_components = [c for c in components if c is not None]
    undetected_points = sum(1 for c in components if c is None)
    total_scaled = sum(scored_components) / len(scored_components) * 7 if scored_components else 0.0

    tier = "Strong" if total_scaled >= 5 else ("Moderate" if total_scaled >= 3 else "Weak")
    return {
        "p1_score": round(p1, 2) if p1 is not None else None,
        "p2_score": round(p2, 2) if p2 is not None else None,
        "p3_score": round(p3, 2) if p3 is not None else None,
        "p4_score": round(p4, 2) if p4 is not None else None,
        "p5_score": round(p5, 2) if p5 is not None else None,
        "p6_score": round(p6, 2) if p6 is not None else None,
        "p7_score": round(p7, 2) if p7 is not None else None,
        "total_score_of_7": round(total_scaled, 2),
        "undetected_points": undetected_points,
        "citability_tier": tier,
    }


FULL_DATA_HEADERS = [
    ("url", "URL"), ("title", "Title"),
    ("p1_has_best_or_top", "P1: Has Best/Top Keyword"),
    ("p1_has_year_in_title", "P1: Has Year in Title"),
    ("p1_has_for_clause", "P1: Has 'For [Persona/Use Case]' Clause"),
    ("p1_buyer_query_match", "P1: Full Buyer Query Match"),
    ("p1_score", "P1 Score (0-1)"),
    ("p2_pick_count", "P2: Pick Count"),
    ("p2_pick_detection_method", "P2: Pick Detection Method"),
    ("p2_pick_count_in_range_5_8", "P2: Pick Count in 5-8 Range"),
    ("p2_best_for_mentions", "P2: 'Best For' Mentions (count)"),
    ("p2_has_best_for_labels", "P2: Has Best-For Labels (>=3)"),
    ("p2_quick_answer_near_top", "P2: Quick-Answer Summary Near Top"),
    ("p2_score", "P2 Score (0-1)"),
    ("p3_methodology_heading_present", "P3: Methodology Heading Present"),
    ("p3_criteria_language_present", "P3: Criteria Language in Body"),
    ("p3_visible_byline_present", "P3: Visible Byline Present"),
    ("p3_visible_date_present", "P3: Visible Last-Updated Date"),
    ("p3_score", "P3 Score (0-1)"),
    ("p4_has_comparison_table", "P4: Has Genuine Comparison Table"),
    ("p4_num_genuine_tables", "P4: Num Genuine Tables"),
    ("p4_toc_rendered_as_table", "P4: Page Also Has ToC-as-Table"),
    ("p4_table_column_match_count_of_8", "P4: Column Keyword Match (of 8)"),
    ("p4_score", "P4 Score (0-1)"),
    ("p5_num_pick_sections", "P5: Num Pick Sections"),
    ("p5_vendor_sections_present", "P5: Vendor Sections Present"),
    ("p5_pct_sections_with_features", "P5: % Sections w/ Features"),
    ("p5_pct_sections_with_limits", "P5: % Sections w/ Limits"),
    ("p5_pct_sections_with_pricing", "P5: % Sections w/ Pricing"),
    ("p5_pct_sections_with_best_for", "P5: % Sections w/ Best-For"),
    ("p5_pct_sections_with_choose_if", "P5: % Sections w/ Choose-If"),
    ("p5_score", "P5 Score (0-1)"),
    ("p6_decision_heading_present", "P6: Decision Heading Present"),
    ("p6_choose_if_avoid_pattern_in_body", "P6: Choose/Avoid Pattern in Body"),
    ("p6_decision_framework_present", "P6: Decision Framework Present"),
    ("p6_score", "P6 Score (0-1)"),
    ("p7_has_stats", "P7: Has Stats/Data Points"),
    ("p7_external_link_count", "P7: External Link Count"),
    ("p7_has_faq_section", "P7: Has Visible FAQ"),
    ("p7_has_faq_schema", "P7: Has FAQPage Schema"),
    ("p7_has_article_schema", "P7: Has Article Schema"),
    ("p7_has_itemlist_schema", "P7: Has ItemList Schema"),
    ("p7_score", "P7 Score (0-1)"),
    ("undetected_points", "Points Undetected (out of 7)"),
    ("total_score_of_7", "TOTAL SCORE (of 7)"),
    ("citability_tier", "Citability Tier"),
]

POINT_JOBS = {
    "1": "Match the question AI is answering",
    "2": "Give AI a safe summary to quote",
    "3": "Make ranking feel earned",
    "4": "Make the market map extractable",
    "5": "Clean entity facts + honest tradeoffs",
    "6": "Answer follow-up buyer questions",
    "7": "Make the page easy to verify and cite",
}
POINT_LABELS = {
    "1": "Buyer Query - title matches 'Best [category] for [persona/use case] in [year]'",
    "2": "Quick Answer - 5-8 picks, best-for labels, 1-line why, near top",
    "3": "Methodology - criteria, sources, author, last updated",
    "4": "Comparison Table - Tool / Best for / Strength / Limit / Pricing",
    "5": "Vendor Sections - per-pick Features / Limits / Pricing / Choose-if",
    "6": "Decision Framework - Choose X if... / Avoid Z if...",
    "7": "Proof + FAQ + Schema - stats, sources, FAQs, Article/ItemList/FAQ schema",
}


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def autosize(ws, ncols, max_width=45):
    for col in range(1, ncols + 1):
        letter = get_column_letter(col)
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 200) + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def pct_str(cond, rows):
    n = len(rows)
    c = sum(1 for r in rows if cond(r))
    return c, round(100 * c / n, 1) if n else 0


def avg_of(key, rows):
    vals = [r[key] for r in rows if r.get(key) is not None]
    return round(statistics.mean(vals), 2) if vals else None


def build_workbook(rows, path):
    n = len(rows)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Full Data"
    ws1.append([h[1] for h in FULL_DATA_HEADERS])
    for row in rows:
        ws1.append([row.get(h[0], "") for h in FULL_DATA_HEADERS])
    style_header(ws1, len(FULL_DATA_HEADERS))
    autosize(ws1, len(FULL_DATA_HEADERS))

    ws2 = wb.create_sheet("Summary")
    ws2.append([f"Citable Listicle Architecture - Summary ({n} listicles audited)"])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])

    tier_counts = Counter(r["citability_tier"] for r in rows)
    ws2.append(["Overall Citability Tier Distribution"])
    ws2["A" + str(ws2.max_row)].font = SECTION_FONT
    ws2.append(["Tier", "Count", "Percent"])
    for tier in ["Strong", "Moderate", "Weak"]:
        c = tier_counts.get(tier, 0)
        ws2.append([tier, c, round(100 * c / n, 1) if n else 0])
    ws2.append([])

    avg_total = avg_of("total_score_of_7", rows)
    ws2.append([f"Average total score: {avg_total} / 7"])
    ws2.append([])

    checks = [
        ("P1 - Has Best/Top keyword in title", lambda r: r["p1_has_best_or_top"]),
        ("P1 - Has year in title", lambda r: r["p1_has_year_in_title"]),
        ("P1 - Has 'for [persona/use case]' clause", lambda r: r["p1_has_for_clause"]),
        ("P1 - Full buyer query match", lambda r: r["p1_buyer_query_match"]),
        ("P2 - Pick count in ideal 5-8 range", lambda r: r["p2_pick_count_in_range_5_8"] is True),
        ("P2 - Has best-for labels (>=3 mentions)", lambda r: r["p2_has_best_for_labels"]),
        ("P2 - Quick-answer summary appears near top", lambda r: r["p2_quick_answer_near_top"]),
        ("P3 - Has methodology heading", lambda r: r["p3_methodology_heading_present"]),
        ("P3 - Has criteria language in body", lambda r: r["p3_criteria_language_present"]),
        ("P3 - Visible byline", lambda r: r["p3_visible_byline_present"]),
        ("P3 - Visible last-updated date", lambda r: r["p3_visible_date_present"]),
        ("P4 - Has genuine comparison table (excl. ToC)", lambda r: r["p4_has_comparison_table"]),
        ("P4 - Table column headers match framework (>=3 of 8 keywords)", lambda r: r["p4_table_column_match_count_of_8"] >= 3),
        ("P5 - Has dedicated vendor sections per pick", lambda r: r["p5_vendor_sections_present"]),
        ("P6 - Has explicit decision framework", lambda r: r["p6_decision_framework_present"]),
        ("P7 - Has stats/data points", lambda r: r["p7_has_stats"]),
        ("P7 - Has visible FAQ", lambda r: r["p7_has_faq_section"]),
        ("P7 - Has FAQPage schema", lambda r: r["p7_has_faq_schema"]),
        ("P7 - Has Article schema", lambda r: r["p7_has_article_schema"]),
        ("P7 - Has ItemList schema", lambda r: r["p7_has_itemlist_schema"]),
    ]
    ws2.append(["Per-Check Pass Rates"])
    ws2["A" + str(ws2.max_row)].font = SECTION_FONT
    ws2.append(["Check", "Count", "Percent"])
    for label, cond in checks:
        c, p = pct_str(cond, rows)
        ws2.append([label, c, p])
    ws2.append([])

    not_detected = sum(1 for r in rows if r["p2_pick_detection_method"] == "not_detected")
    if not_detected:
        ws2.append([f"Pick structure not detected on {not_detected}/{n} posts - Points 2 and 5 excluded from their score, not penalized"])
        ws2.append([])

    worst = sorted(rows, key=lambda r: r["total_score_of_7"])[:15]
    ws2.append(["Lowest-Scoring Listicles (highest priority to fix)"])
    ws2["A" + str(ws2.max_row)].font = SECTION_FONT
    ws2.append(["URL", "Total Score (of 7)", "Tier"])
    for r in worst:
        ws2.append([r["url"], r["total_score_of_7"], r["citability_tier"]])
    ws2.append([])

    best = sorted(rows, key=lambda r: r["total_score_of_7"], reverse=True)[:15]
    ws2.append(["Highest-Scoring Listicles (models to replicate)"])
    ws2["A" + str(ws2.max_row)].font = SECTION_FONT
    ws2.append(["URL", "Total Score (of 7)", "Tier"])
    for r in best:
        ws2.append([r["url"], r["total_score_of_7"], r["citability_tier"]])
    autosize(ws2, 3, max_width=70)

    ws3 = wb.create_sheet("Point Detail")
    ws3.append(["#", "Point", "Sitewide Result", "Job"])
    style_header(ws3, 4)
    c1, p1v = pct_str(lambda r: r["p1_buyer_query_match"], rows)
    c4, p4v = pct_str(lambda r: r["p4_has_comparison_table"], rows)
    c5, p5v = pct_str(lambda r: r["p5_vendor_sections_present"], rows)
    c6, p6v = pct_str(lambda r: r["p6_decision_framework_present"], rows)
    c_qat, p_qat = pct_str(lambda r: r["p2_quick_answer_near_top"], rows)
    c_byline, p_byline = pct_str(lambda r: r["p3_visible_byline_present"], rows)
    c_date, p_date = pct_str(lambda r: r["p3_visible_date_present"], rows)
    results = {
        "1": f"{c1}/{n} ({p1v}%) fully match 'Best/Top X for Y'",
        "2": f"Avg score {avg_of('p2_score', rows)}/1 where detectable; {c_qat}/{n} ({p_qat}%) have a quick-answer summary near the top",
        "3": f"Avg score {avg_of('p3_score', rows)}/1; visible byline {p_byline}%, visible date {p_date}%",
        "4": f"{c4}/{n} ({p4v}%) have a genuine comparison table",
        "5": f"{c5}/{n} ({p5v}%) have dedicated per-pick sections",
        "6": f"{c6}/{n} ({p6v}%) have an explicit decision framework",
        "7": f"Avg proof score {avg_of('p7_proof_score_of_6', rows)}/6",
    }
    for num in "1234567":
        ws3.append([num, POINT_LABELS[num], results[num], POINT_JOBS[num]])
    autosize(ws3, 4, max_width=80)
    for row_cells in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)


def build_report(rows, path):
    n = len(rows)
    tiers = Counter(r["citability_tier"] for r in rows)
    avg_total = avg_of("total_score_of_7", rows)

    def cp(cond):
        return pct_str(cond, rows)

    lines = []
    lines.append("# Citable Listicle Architecture Audit")
    lines.append("")
    lines.append(f"Covers {n} listicle URLs, checked against the 7-point Citable Listicle Architecture "
                 "framework (buyer query, quick answer, methodology, comparison table, vendor sections, "
                 "decision framework, proof+FAQ+schema).")
    lines.append("")
    lines.append(f"**Overall: average score is {avg_total}/7** "
                 f"({tiers.get('Weak',0)} Weak, {tiers.get('Moderate',0)} Moderate, {tiers.get('Strong',0)} Strong).")
    lines.append("")
    not_detected = sum(1 for r in rows if r["p2_pick_detection_method"] == "not_detected")
    if not_detected:
        pct_nd = round(100 * not_detected / n, 1)
        lines.append(f"**Note:** pick structure could not be detected on {not_detected}/{n} posts ({pct_nd}%) - "
                     "Points 2 and 5 were excluded from those posts' scores rather than penalized. If this is a "
                     "large share, the site may use a listicle format this script doesn't recognize (not numbered "
                     "headings or ordered lists) - spot-check a few of those URLs in a browser.")
        lines.append("")
    lines.append("## The 7 points")
    lines.append("")
    lines.append("| # | Point | Result | Job |")
    lines.append("|---|---|---|---|")
    c1, p1v = cp(lambda r: r["p1_buyer_query_match"])
    c_best, p_best = cp(lambda r: r["p1_has_best_or_top"])
    c_for, p_for = cp(lambda r: r["p1_has_for_clause"])
    lines.append(f"| 1 | Buyer Query | {c_best}/{n} ({p_best}%) use Best/Top in the title, but only {c_for}/{n} ({p_for}%) add a persona/use-case clause - {c1}/{n} ({p1v}%) fully match | Match the question AI is answering |")
    c_range, p_range = cp(lambda r: r["p2_pick_count_in_range_5_8"] is True)
    c_qat, p_qat = cp(lambda r: r["p2_quick_answer_near_top"])
    lines.append(f"| 2 | Quick Answer | {c_range}/{n} ({p_range}%) keep picks to the ideal 5-8 range, {c_qat}/{n} ({p_qat}%) have a scannable summary before the first heading | Give AI a safe summary to quote |")
    c_meth, p_meth = cp(lambda r: r["p3_methodology_heading_present"])
    c_byline, p_byline = cp(lambda r: r["p3_visible_byline_present"])
    c_date, p_date = cp(lambda r: r["p3_visible_date_present"])
    lines.append(f"| 3 | Methodology | {c_meth}/{n} ({p_meth}%) have a methodology heading; visible byline {p_byline}%, visible date {p_date}% | Make ranking feel earned |")
    c_tbl, p_tbl = cp(lambda r: r["p4_has_comparison_table"])
    avg_col = avg_of("p4_table_column_match_count_of_8", rows)
    lines.append(f"| 4 | Comparison Table | {c_tbl}/{n} ({p_tbl}%) have a genuine comparison table (Table-of-Contents tables excluded); avg column-keyword match {avg_col}/8 | Make the market map extractable |")
    c_vs, p_vs = cp(lambda r: r["p5_vendor_sections_present"])
    lines.append(f"| 5 | Vendor Sections | {c_vs}/{n} ({p_vs}%) have a dedicated section per pick; avg choose-if language in {avg_of('p5_pct_sections_with_choose_if', rows)}% of sections | Clean entity facts + honest tradeoffs |")
    c_df, p_df = cp(lambda r: r["p6_decision_framework_present"])
    lines.append(f"| 6 | Decision Framework | {c_df}/{n} ({p_df}%) have an explicit decision framework | Answer follow-up buyer questions |")
    avg_proof = avg_of("p7_proof_score_of_6", rows)
    c_il, p_il = cp(lambda r: r["p7_has_itemlist_schema"])
    lines.append(f"| 7 | Proof + FAQ + Schema | Avg {avg_proof}/6; ItemList schema present on {c_il}/{n} ({p_il}%) | Make the page easy to verify and cite |")
    lines.append("")
    lines.append("## What this means, in priority order")
    lines.append("")
    lines.append("Read the per-check pass rates in the Summary tab and rank fixes by: (a) cheapest to ship "
                 "sitewide (schema/template changes) before (b) per-post content rewrites. Typical priority "
                 "order based on this framework:")
    lines.append("")
    lines.append("1. Any point at or near 0% that is a template-level fix (schema markup, byline/date display) "
                 "- fix once, apply everywhere.")
    lines.append("2. Points with a genuine structural gap on most posts (e.g. no methodology section, no "
                 "decision framework) - these need a reusable content pattern, not a full rewrite.")
    lines.append("3. Points that are close but inconsistent (e.g. comparison tables exist but use different "
                 "column vocabulary) - standardize the template rather than rebuilding from scratch.")
    lines.append("")
    lines.append("## Data notes")
    lines.append("")
    lines.append("- Table detection distinguishes genuine comparison tables from Table-of-Contents blocks "
                 "rendered as `<table>` elements - do not trust a raw `<table>` count without this filter.")
    lines.append("- Visible-byline detection is text-pattern only (no CSS-class heuristic) - it can miss a "
                 "real byline written in an unusual format, and may over- or under-count depending on site copy.")
    lines.append("- Schema (Article/FAQPage/ItemList) can only be verified by parsing raw HTML - a page-summary "
                 "tool that strips `<script>` tags will silently report 0% schema even when it exists.")
    lines.append("- Per-point scores are a 0-1 composite of that point's sub-checks; sub-checks that could not "
                 "be evaluated are excluded from the average rather than counted as failing.")
    lines.append("")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def main():
    if len(sys.argv) < 2:
        print("Usage: python build_outputs.py <extract_json> [output_prefix]")
        sys.exit(1)

    extract_path = sys.argv[1]
    prefix = sys.argv[2] if len(sys.argv) > 2 else "listicle_audit"

    with open(extract_path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    ok_rows = [r for r in raw if r.get("fetch_status") == "ok"]
    error_rows = [r for r in raw if r.get("fetch_status") != "ok"]
    if error_rows:
        print(f"[WARNING] {len(error_rows)} URLs failed to fetch and are excluded from scoring:")
        for r in error_rows[:10]:
            print(f"   {r['url']} - {r.get('fetch_status')}")

    rows = []
    for r in ok_rows:
        scores = compute_scores(r)
        merged = dict(r)
        merged.update(scores)
        rows.append(merged)

    build_workbook(rows, f"{prefix}_workbook.xlsx")
    build_report(rows, f"{prefix}_report.md")
    print(f"[DONE] {len(rows)} listicles scored. Wrote {prefix}_workbook.xlsx and {prefix}_report.md")


if __name__ == "__main__":
    main()
