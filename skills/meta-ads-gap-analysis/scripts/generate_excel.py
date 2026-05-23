"""
Phase 5 - Generate final Excel report and markdown summary.
Run after parse_video_analyses.py has been executed.

Usage:
    python generate_excel.py <analysis_name>

Reads config from outputs/<analysis_name>/config.json
Input:  outputs/<analysis_name>/analysis/*.json
Output: outputs/<analysis_name>/<analysis_name>_analysis.xlsx
        outputs/<analysis_name>/<analysis_name>_summary.md
"""

import json
import sys
from pathlib import Path
from collections import Counter
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = Path.cwd()


def load_config(analysis_name):
    config_path = BASE / "outputs" / analysis_name / "config.json"
    if not config_path.exists():
        print(f"ERROR: No config.json at {config_path}")
        sys.exit(1)
    return json.loads(config_path.read_text(encoding="utf-8"))


def load_json(analysis_dir, filename):
    p = analysis_dir / "analysis" / filename
    if not p.exists():
        print(f"[WARN] {p} not found")
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


# -- Styling helpers ----------------------------------------------------------

COLOR_A = "FF4A90E2"      # blue for brand_a
COLOR_B = "FF27AE60"      # green for brand_b
HEADER_COLOR = "FF2C3E50" # dark main headers
ACCENT_COLOR = "FFEAF4FB"


def header_style(ws, row, col, value, bg=HEADER_COLOR, font_color="FFFFFFFF", bold=True, fontsize=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_color, size=fontsize)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_df(ws, df, start_row=1, header_bg=HEADER_COLOR):
    for col_idx, col_name in enumerate(df.columns, 1):
        header_style(ws, start_row, col_idx, col_name, bg=header_bg)
    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    return start_row + len(df) + 1


# -- Tab builders -------------------------------------------------------------

def build_overview(ws, meta_a, meta_b, video_stats, label_a, label_b):
    ws.title = "Overview"

    header_style(ws, 1, 1, f"{label_a} vs {label_b} - Meta Ads Competitive Analysis", bg=HEADER_COLOR, fontsize=14)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 30

    header_style(ws, 3, 1, "Dimension", bg="FF34495E")
    header_style(ws, 3, 2, label_a, bg=COLOR_A)
    header_style(ws, 3, 3, label_b, bg=COLOR_B)
    header_style(ws, 3, 4, "Gap / Finding", bg=HEADER_COLOR)

    s_a = video_stats.get(meta_a.get("brand", ""), {})
    s_b = video_stats.get(meta_b.get("brand", ""), {})
    fmt_a = meta_a.get("format_breakdown", {})
    fmt_b = meta_b.get("format_breakdown", {})

    def fmt_str(d):
        return ", ".join(f"{k}: {v}" for k, v in sorted(d.items(), key=lambda x: -x[1]))

    rows = [
        ("Total Active Ads",
         str(meta_a.get("total_ads", "N/A")),
         str(meta_b.get("total_ads", "N/A")),
         "Compare ad volume as proxy for spend/investment"),
        ("Format Mix",
         fmt_str(fmt_a),
         fmt_str(fmt_b),
         "Compare reliance on video vs catalog/DCO formats"),
        ("Influencer / Branded Content",
         f"{meta_a.get('branded_content_count', 0)} ads | pages: {', '.join(meta_a.get('branded_content_pages', {}).keys()) or 'none'}",
         f"{meta_b.get('branded_content_count', 0)} ads | pages: {', '.join(meta_b.get('branded_content_pages', {}).keys()) or 'none'}",
         "Influencer partnerships drive trust and reach"),
        ("Avg New Ads/Week",
         str(meta_a.get("avg_ads_per_week", "N/A")),
         str(meta_b.get("avg_ads_per_week", "N/A")),
         "Creative velocity / refresh rate"),
        ("Ads with Offers",
         str(meta_a.get("offer_count", 0)),
         str(meta_b.get("offer_count", 0)),
         "Bundle/discount ad share"),
        ("AI-Generated Content",
         str(meta_a.get("ai_generated_count", 0)),
         str(meta_b.get("ai_generated_count", 0)),
         "contains_digital_created_media flag"),
        ("Emotional Hook Style",
         _fmt_dict(s_a.get("emotional_angles", {})),
         _fmt_dict(s_b.get("emotional_angles", {})),
         "Dominant emotional angles from video analysis"),
        ("Narrator Type",
         _fmt_dict(s_a.get("narrator_types", {})),
         _fmt_dict(s_b.get("narrator_types", {})),
         "Who delivers the message on screen"),
        ("Production Quality",
         _fmt_dict(s_a.get("production_quality", {})),
         _fmt_dict(s_b.get("production_quality", {})),
         ""),
    ]

    for i, (dim, val_a, val_b, gap) in enumerate(rows, 4):
        ws.cell(row=i, column=1, value=dim).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val_a)
        ws.cell(row=i, column=3, value=val_b)
        ws.cell(row=i, column=4, value=gap)
        for col in range(1, 5):
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=ACCENT_COLOR)

    set_col_widths(ws, [30, 35, 40, 45])


def _fmt_dict(d):
    if not d:
        return "N/A"
    return ", ".join(f"{k}: {v}" for k, v in sorted(d.items(), key=lambda x: -x[1]))


def build_ads_tab(ws, slim_rows, tab_label, bg_color):
    ws.title = tab_label

    df_rows = []
    for r in slim_rows:
        df_rows.append({
            "Ad ID": r.get("ad_archive_id", ""),
            "Start Date": r.get("start_date_formatted", ""),
            "Format": r.get("snapshot/display_format", ""),
            "Title": r.get("snapshot/title", "")[:100],
            "Body (first 200c)": r.get("snapshot/body/text", "")[:200],
            "Caption": r.get("snapshot/caption", "")[:100],
            "CTA": r.get("snapshot/cta_type", ""),
            "Link URL": r.get("snapshot/link_url", "")[:80],
            "Branded Content Page": r.get("snapshot/branded_content/page_name", ""),
            "AI Generated": r.get("contains_digital_created_media", ""),
        })

    df = pd.DataFrame(df_rows)
    write_df(ws, df, header_bg=bg_color)
    set_col_widths(ws, [20, 12, 8, 40, 50, 30, 12, 45, 25, 12])


def build_hooks_tab(ws, video_a, video_b, label_a, label_b):
    ws.title = "Hooks Comparison"
    headers = ["Brand", "Ad ID", "Emotional Angle", "Narrator Type", "Hook (0-3s)", "Key Claim/USP", "Production Quality", "Offers/Pricing"]
    for col_idx, h in enumerate(headers, 1):
        header_style(ws, 1, col_idx, h, bg=HEADER_COLOR)

    row = 2
    for entry in video_a:
        ws.cell(row=row, column=1, value=label_a)
        ws.cell(row=row, column=2, value=entry.get("ad_id", ""))
        ws.cell(row=row, column=3, value=entry.get("emotional_angle", ""))
        ws.cell(row=row, column=4, value=entry.get("narrator_type", ""))
        ws.cell(row=row, column=5, value=entry.get("hook", "")[:200])
        ws.cell(row=row, column=6, value=entry.get("key_claim", "")[:150])
        ws.cell(row=row, column=7, value=entry.get("production_quality", ""))
        ws.cell(row=row, column=8, value=entry.get("offers_pricing", "")[:100])
        for col in range(1, 9):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    for entry in video_b:
        ws.cell(row=row, column=1, value=label_b)
        ws.cell(row=row, column=2, value=entry.get("ad_id", ""))
        ws.cell(row=row, column=3, value=entry.get("emotional_angle", ""))
        ws.cell(row=row, column=4, value=entry.get("narrator_type", ""))
        ws.cell(row=row, column=5, value=entry.get("hook", "")[:200])
        ws.cell(row=row, column=6, value=entry.get("key_claim", "")[:150])
        ws.cell(row=row, column=7, value=entry.get("production_quality", ""))
        ws.cell(row=row, column=8, value=entry.get("offers_pricing", "")[:100])
        for col in range(1, 9):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    set_col_widths(ws, [14, 20, 18, 18, 55, 45, 18, 30])


def build_product_focus_tab(ws, meta_a, meta_b, label_a, label_b):
    ws.title = "Product Focus"

    header_style(ws, 1, 1, f"{label_a} - Top Destination URLs", bg=COLOR_A, fontsize=12)
    ws.merge_cells("A1:B1")
    header_style(ws, 2, 1, "URL Path", bg=COLOR_A)
    header_style(ws, 2, 2, "Ad Count", bg=COLOR_A)

    row = 3
    for url, count in (meta_a.get("product_focus") or {}).items():
        ws.cell(row=row, column=1, value=url)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    header_style(ws, row, 1, f"{label_b} - Top Destination URLs", bg=COLOR_B, fontsize=12)
    ws.merge_cells(f"A{row}:B{row}")
    row += 1
    header_style(ws, row, 1, "URL Path", bg=COLOR_B)
    header_style(ws, row, 2, "Ad Count", bg=COLOR_B)
    row += 1

    for url, count in (meta_b.get("product_focus") or {}).items():
        ws.cell(row=row, column=1, value=url)
        ws.cell(row=row, column=2, value=count)
        row += 1

    set_col_widths(ws, [60, 12])


def build_influencer_tab(ws, meta_a, meta_b, label_a, label_b):
    ws.title = "Influencer Analysis"

    branded_a = meta_a.get("branded_content_ads", [])
    branded_b = meta_b.get("branded_content_ads", [])
    count_a = meta_a.get("branded_content_count", 0)
    count_b = meta_b.get("branded_content_count", 0)

    row = 1
    header_style(ws, row, 1, f"{label_a} - {count_a} Branded Content Ads", bg=COLOR_A, fontsize=12)
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    if branded_a:
        df_rows = [{"Ad ID": r.get("ad_archive_id", ""), "Start Date": r.get("start_date_formatted", ""),
                    "Format": r.get("snapshot/display_format", ""), "Influencer Page": r.get("snapshot/branded_content/page_name", ""),
                    "Title": r.get("snapshot/title", "")[:80], "Body Copy": r.get("snapshot/body/text", "")[:250]}
                   for r in branded_a]
        write_df(ws, pd.DataFrame(df_rows), start_row=row, header_bg=COLOR_A)
        row += len(branded_a) + 2
    else:
        ws.cell(row=row, column=1, value=f"No branded content ads for {label_a}")
        row += 2

    header_style(ws, row, 1, f"{label_b} - {count_b} Branded Content Ads", bg=COLOR_B, fontsize=12)
    ws.merge_cells(f"A{row}:F{row}")
    row += 1

    if branded_b:
        df_rows = [{"Ad ID": r.get("ad_archive_id", ""), "Start Date": r.get("start_date_formatted", ""),
                    "Format": r.get("snapshot/display_format", ""), "Influencer Page": r.get("snapshot/branded_content/page_name", ""),
                    "Title": r.get("snapshot/title", "")[:80], "Body Copy": r.get("snapshot/body/text", "")[:250]}
                   for r in branded_b]
        write_df(ws, pd.DataFrame(df_rows), start_row=row, header_bg=COLOR_B)
    else:
        ws.cell(row=row, column=1, value=f"No branded content ads for {label_b}")

    set_col_widths(ws, [20, 12, 8, 25, 45, 60])


def build_gap_analysis_tab(ws, s_a, s_b, meta_a, meta_b, label_a, label_b):
    ws.title = "Gap Analysis"

    header_style(ws, 1, 1, "Gap Dimension", bg=HEADER_COLOR)
    header_style(ws, 1, 2, label_a, bg=COLOR_A)
    header_style(ws, 1, 3, label_b, bg=COLOR_B)
    header_style(ws, 1, 4, "Opportunity", bg="FF8E44AD", font_color="FFFFFFFF")
    header_style(ws, 1, 5, "Priority", bg=HEADER_COLOR)

    fmt_a = meta_a.get("format_breakdown", {})
    fmt_b = meta_b.get("format_breakdown", {})
    angles_a = s_a.get("emotional_angles", {})
    angles_b = s_b.get("emotional_angles", {})
    narrators_a = s_a.get("narrator_types", {})
    narrators_b = s_b.get("narrator_types", {})
    inf_a = meta_a.get("branded_content_count", 0)
    inf_b = meta_b.get("branded_content_count", 0)
    offers_a = meta_a.get("offer_count", 0)
    offers_b = meta_b.get("offer_count", 0)
    total_a = meta_a.get("total_ads", 1)
    total_b = meta_b.get("total_ads", 1)

    gaps = [
        (
            "1. Format Strategy",
            _fmt_dict(fmt_a),
            _fmt_dict(fmt_b),
            "Identify which brand relies more on catalog/DCO vs human-first video. Human-first video typically drives stronger consideration.",
            "HIGH"
        ),
        (
            "2. Influencer / Branded Content",
            f"{inf_a} ads ({round(inf_a/total_a*100)}%) via influencer pages",
            f"{inf_b} ads ({round(inf_b/total_b*100)}%) via influencer pages",
            "The brand with fewer influencer ads should consider adding micro-influencers to build social proof and authentic reach.",
            "CRITICAL" if abs(inf_a - inf_b) > 5 else "MEDIUM"
        ),
        (
            "3. Emotional Hook Strategy",
            _fmt_dict(angles_a),
            _fmt_dict(angles_b),
            "Brand with more product/feature hooks should shift toward identity/aspiration/fear-of-harm emotional angles to drive consideration.",
            "HIGH"
        ),
        (
            "4. Narrator Type",
            _fmt_dict(narrators_a),
            _fmt_dict(narrators_b),
            "On-camera human narrators outperform text-only and VO in trust-building. Shift toward more human faces on screen.",
            "HIGH"
        ),
        (
            "5. Offer / Bundle Mechanics",
            f"{offers_a} of {total_a} ads ({round(offers_a/total_a*100)}%) feature offers/pricing",
            f"{offers_b} of {total_b} ads ({round(offers_b/total_b*100)}%) feature offers/pricing",
            "Brand with fewer offer ads should create bundle/kit-first creatives with explicit pricing to compete on value perception.",
            "MEDIUM"
        ),
        (
            "6. Carousel Format",
            "No carousel ads (0)",
            "No carousel ads (0)",
            "UNTAPPED BY BOTH BRANDS. Carousel allows multi-step storytelling (problem > ingredient > proof > CTA). Neither brand is testing this.",
            "MEDIUM"
        ),
        (
            "7. Creative Velocity",
            f"~{meta_a.get('avg_ads_per_week', 'N/A')} new ads/week",
            f"~{meta_b.get('avg_ads_per_week', 'N/A')} new ads/week",
            "Brand with lower velocity should increase creative refresh. Ad fatigue sets in after 4-6 weeks. Aim for 6-8 new creatives/month.",
            "HIGH"
        ),
    ]

    for i, (dim, val_a, val_b, opportunity, priority) in enumerate(gaps, 2):
        ws.cell(row=i, column=1, value=dim).font = Font(bold=True)
        ws.cell(row=i, column=2, value=val_a)
        ws.cell(row=i, column=3, value=val_b)
        ws.cell(row=i, column=4, value=opportunity)
        ws.cell(row=i, column=5, value=priority)
        if priority == "CRITICAL":
            ws.cell(row=i, column=5).font = Font(bold=True, color="FFCC0000")
        elif priority == "HIGH":
            ws.cell(row=i, column=5).font = Font(bold=True, color="FFE67E22")
        for col in range(1, 6):
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=ACCENT_COLOR)

    set_col_widths(ws, [30, 40, 40, 55, 10])


def build_recommendations_tab(ws, meta_a, meta_b, s_a, s_b, label_a, label_b):
    ws.title = "Recommendations"

    # Determine which brand is "ours" (brand_b) and "theirs" (brand_a)
    inf_a = meta_a.get("branded_content_count", 0)
    inf_b = meta_b.get("branded_content_count", 0)
    offers_a = meta_a.get("offer_count", 0)
    offers_b = meta_b.get("offer_count", 0)
    total_b = meta_b.get("total_ads", 1)
    vel_a = meta_a.get("avg_ads_per_week", 0) or 0
    vel_b = meta_b.get("avg_ads_per_week", 0) or 0

    header_style(ws, 1, 1, f"{label_b} - Actionable Recommendations (based on {label_a} comparison)", bg=COLOR_B, fontsize=13)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 30

    header_style(ws, 2, 1, "#", bg=HEADER_COLOR)
    header_style(ws, 2, 2, "Recommendation", bg=HEADER_COLOR)
    header_style(ws, 2, 3, f"Evidence from {label_a}", bg=HEADER_COLOR)
    header_style(ws, 2, 4, "Execution Steps", bg=HEADER_COLOR)

    recs = []

    if inf_a > inf_b + 3:
        recs.append((
            f"[CRITICAL] Launch an influencer / branded content program",
            f"{label_a} runs {inf_a} ads ({round(inf_a/meta_a.get('total_ads',1)*100)}% of library) via influencer pages. {label_b} has {inf_b}. Influencer content builds trust that brand-voice ads cannot replicate.",
            f"1. Identify 3-5 micro-influencers (50K-500K followers) relevant to your category.\n2. Brief them on your core USP.\n3. Create 30-60s UGC-style videos showing product in real use.\n4. Run as branded content (whitelisted) alongside your own ads.",
        ))

    recs.append((
        "[HIGH] Lead videos with emotional identity/aspiration hooks",
        f"{label_a}'s top emotional angles: {_fmt_dict(s_a.get('emotional_angles', {}))}. {label_b}'s: {_fmt_dict(s_b.get('emotional_angles', {}))}.",
        "1. Rewrite video hooks to open with a relatable emotional moment (not a product shot).\n2. Test 3 hook variants: problem-first, aspiration-first, social-proof-first.\n3. Use Meta A/B split test to identify the winning hook type.",
    ))

    if offers_a / max(meta_a.get("total_ads", 1), 1) > offers_b / max(total_b, 1) + 0.2:
        recs.append((
            "[HIGH] Create bundle/kit-first ad creatives with explicit pricing",
            f"{label_a} runs offers in {round(offers_a/meta_a.get('total_ads',1)*100)}% of ads (vs {round(offers_b/total_b*100)}% for {label_b}). Bundle ads drive higher AOV.",
            "1. Create a hero bundle landing page combining 3-4 complementary SKUs.\n2. Film 4-6 video ads showing the complete routine.\n3. Include explicit pricing in the creative itself.\n4. Test bundle ads against single-product ads to measure AOV lift.",
        ))

    if vel_a > vel_b + 2:
        recs.append((
            "[HIGH] Increase creative output velocity",
            f"{label_a} launches ~{vel_a} new ads/week vs {label_b}'s ~{vel_b}/week. Creative fatigue sets in after 4-6 weeks of heavy delivery.",
            "1. Set a target of 6-8 new video creatives per month minimum.\n2. Run a monthly creative sprint: shoot 10-15 raw clips, edit into multiple 15-30s ads.\n3. Use existing catalog/DCO for scale while funding new video creative tests.",
        ))

    recs.append((
        "[MEDIUM] Test carousel ad format",
        "Neither brand currently runs carousel ads. Carousel allows multi-step storytelling: problem > ingredient > before/after > proof > CTA.",
        "1. Create a 5-slide carousel with sequential storytelling.\n2. Test with a small daily budget for 7 days.\n3. Compare CTR and CVR vs single-image and video formats.",
    ))

    recs.append((
        "[MEDIUM] Add LEARN_MORE consideration-stage ads",
        "Most ads use SHOP_NOW (bottom funnel only). Adding educational mid-funnel ads targeting warm audiences improves overall conversion.",
        "1. Create 3-5 educational video ads (30-45s) linking to ingredient/safety content.\n2. Use LEARN_MORE CTA.\n3. Target warm audiences: website visitors, add-to-cart abandoners.",
    ))

    recs.append((
        "[MEDIUM] Plan seasonal sale campaigns",
        f"{label_a} runs themed sale events with explicit pricing in creative. Creates urgency and price clarity.",
        "1. Plan 2-3 seasonal campaigns per quarter with a dedicated visual identity.\n2. Include explicit pricing in the video/image creative itself.\n3. Run for 7-10 days with dedicated landing pages.",
    ))

    for i, (rec, evidence, steps) in enumerate(recs, 3):
        ws.cell(row=i, column=1, value=str(i-2)).font = Font(bold=True)
        ws.cell(row=i, column=2, value=rec)
        ws.cell(row=i, column=3, value=evidence)
        ws.cell(row=i, column=4, value=steps)
        ws.row_dimensions[i].height = 90
        for col in range(1, 5):
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        if i % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=ACCENT_COLOR)

    set_col_widths(ws, [5, 35, 50, 60])


# -- Markdown report ----------------------------------------------------------

def generate_markdown(meta_a, meta_b, s_a, s_b, label_a, label_b, analysis_name):
    fmt_a = meta_a.get("format_breakdown", {})
    fmt_b = meta_b.get("format_breakdown", {})

    lines = [
        f"# {label_a} vs {label_b} - Meta Ads Competitive Analysis",
        f"_Analysis: {analysis_name} | {label_a}: {meta_a.get('total_ads', 0)} ads | {label_b}: {meta_b.get('total_ads', 0)} ads_",
        "",
        "---",
        "",
        "## Executive Summary",
        "",
        f"This report compares Meta Ads Library data for {label_a} and {label_b}. "
        f"{label_a} is running {meta_a.get('total_ads', 0)} active ads vs {label_b}'s {meta_b.get('total_ads', 0)}. "
        f"Key gaps: influencer use ({meta_a.get('branded_content_count', 0)} vs {meta_b.get('branded_content_count', 0)} branded content ads), "
        f"format strategy ({_fmt_dict(fmt_a)} vs {_fmt_dict(fmt_b)}), "
        f"and offer mechanics ({meta_a.get('offer_count', 0)} vs {meta_b.get('offer_count', 0)} ads with pricing/bundles).",
        "",
        "---",
        "",
        "## Key Findings",
        "",
        "### 1. Format Strategy",
        "",
        f"- **{label_a}**: {_fmt_dict(fmt_a)}",
        f"- **{label_b}**: {_fmt_dict(fmt_b)}",
        "",
        "### 2. Influencer / Branded Content",
        "",
        f"- **{label_a}**: {meta_a.get('branded_content_count', 0)} branded content ads via: {', '.join(meta_a.get('branded_content_pages', {}).keys()) or 'none'}",
        f"- **{label_b}**: {meta_b.get('branded_content_count', 0)} branded content ads via: {', '.join(meta_b.get('branded_content_pages', {}).keys()) or 'none'}",
        "",
        "### 3. Emotional Hook Analysis (video analysis)",
        "",
        f"**{label_a} ({s_a.get('total', 0)} analyzed):**",
        f"- Emotional angles: {_fmt_dict(s_a.get('emotional_angles', {}))}",
        f"- Narrator types: {_fmt_dict(s_a.get('narrator_types', {}))}",
        f"- Production quality: {_fmt_dict(s_a.get('production_quality', {}))}",
        "",
        f"**{label_b} ({s_b.get('total', 0)} analyzed):**",
        f"- Emotional angles: {_fmt_dict(s_b.get('emotional_angles', {}))}",
        f"- Narrator types: {_fmt_dict(s_b.get('narrator_types', {}))}",
        f"- Production quality: {_fmt_dict(s_b.get('production_quality', {}))}",
        "",
        "### 4. Offer Mechanics",
        "",
        f"- **{label_a}**: {meta_a.get('offer_count', 0)}/{meta_a.get('total_ads', 0)} ads feature explicit offers or pricing",
        f"- **{label_b}**: {meta_b.get('offer_count', 0)}/{meta_b.get('total_ads', 0)} ads feature explicit offers or pricing",
        "",
        "### 5. Creative Velocity",
        "",
        "_Counts only ads with a start date in the last 28 days, divided by 4 weeks. Evergreen ads started before this window are excluded._",
        "",
        f"- **{label_a}**: ~{meta_a.get('avg_ads_per_week', 'N/A')} new ads/week",
        f"- **{label_b}**: ~{meta_b.get('avg_ads_per_week', 'N/A')} new ads/week",
        "",
        "---",
        "",
        "## Data Sources",
        "",
        f"- Analysis name: {analysis_name}",
        f"- {label_a}: {meta_a.get('total_ads', 0)} active ads",
        f"- {label_b}: {meta_b.get('total_ads', 0)} active ads",
        "- Video analysis via Google Gemini (gemini-3-flash-preview)",
        "- No performance data (reach/impressions/spend) available from Meta Ads Library",
        "",
    ]

    return "\n".join(lines)


# -- Main ---------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_excel.py <analysis_name>")
        sys.exit(1)

    analysis_name = sys.argv[1]
    cfg = load_config(analysis_name)
    analysis_dir = BASE / "outputs" / analysis_name

    label_a = cfg["brand_a"]["label"]
    label_b = cfg["brand_b"]["label"]
    folder_a = cfg["brand_a"]["folder"]
    folder_b = cfg["brand_b"]["folder"]

    print("Loading analysis data...")
    meta_a = load_json(analysis_dir, f"{folder_a}_analysis.json")
    meta_b = load_json(analysis_dir, f"{folder_b}_analysis.json")
    video_data = load_json(analysis_dir, "video_analyses_parsed.json")

    s_a = video_data.get("stats", {}).get(folder_a, {})
    s_b = video_data.get("stats", {}).get(folder_b, {})
    video_a = video_data.get(folder_a, [])
    video_b = video_data.get(folder_b, [])

    wb = Workbook()

    ws1 = wb.active
    build_overview(ws1, meta_a, meta_b, video_data.get("stats", {}), label_a, label_b)

    ws2 = wb.create_sheet("Gap Analysis")
    build_gap_analysis_tab(ws2, s_a, s_b, meta_a, meta_b, label_a, label_b)

    ws3 = wb.create_sheet("Recommendations")
    build_recommendations_tab(ws3, meta_a, meta_b, s_a, s_b, label_a, label_b)

    ws4 = wb.create_sheet("Product Focus")
    build_product_focus_tab(ws4, meta_a, meta_b, label_a, label_b)

    ws5 = wb.create_sheet("Hooks Comparison")
    build_hooks_tab(ws5, video_a, video_b, label_a, label_b)

    # Audio Hooks tab (tab 6) is added by add_audio_hooks_tab.py at index 5

    ws6 = wb.create_sheet("Influencer Analysis")
    build_influencer_tab(ws6, meta_a, meta_b, label_a, label_b)

    ws7 = wb.create_sheet(f"{label_b} Ads")
    build_ads_tab(ws7, meta_b.get("slim_rows", []), f"{label_b} Ads", COLOR_B)

    ws8 = wb.create_sheet(f"{label_a} Ads")
    build_ads_tab(ws8, meta_a.get("slim_rows", []), f"{label_a} Ads", COLOR_A)

    out_path = analysis_dir / f"{analysis_name}_analysis.xlsx"
    wb.save(str(out_path))
    print(f"Excel saved: {out_path}")

    md = generate_markdown(meta_a, meta_b, s_a, s_b, label_a, label_b, analysis_name)
    md_path = analysis_dir / f"{analysis_name}_summary.md"
    md_path.write_text(md, encoding="utf-8")
    print(f"Markdown saved: {md_path}")


if __name__ == "__main__":
    main()
