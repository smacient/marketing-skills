"""
Add an Audio Hooks tab to an existing analysis Excel file.
Run after analyze_audio_hooks_batch.py has processed all videos.

Usage:
    python add_audio_hooks_tab.py <analysis_name>

Reads config from outputs/<analysis_name>/config.json
Input:  outputs/<analysis_name>/audio_hooks/<brand>/<id>.md
Output: Adds/replaces 'Audio Hooks' tab in outputs/<analysis_name>/<analysis_name>_analysis.xlsx
"""

import sys
import json
from pathlib import Path
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = Path.cwd()
FIELDS = ["First Spoken Line", "First On-Screen Text", "Hook Type", "Hook Summary"]


def load_config(analysis_name):
    config_path = BASE / "outputs" / analysis_name / "config.json"
    if not config_path.exists():
        print(f"ERROR: No config.json at {config_path}")
        sys.exit(1)
    return json.loads(config_path.read_text(encoding="utf-8"))


def extract_hook_fields(content):
    result = {f: "" for f in FIELDS}
    lines = content.splitlines()
    current_field = None
    current_value_lines = []

    for line in lines:
        stripped = line.strip()
        matched_field = None

        for field in FIELDS:
            marker = "**" + field + ":**"
            if stripped.startswith(marker):
                matched_field = field
                remainder = stripped[len(marker):].strip()
                break

        if matched_field:
            if current_field:
                result[current_field] = "\n".join(current_value_lines).strip()
            current_field = matched_field
            current_value_lines = [remainder] if remainder else []
        elif current_field:
            current_value_lines.append(stripped)

    if current_field:
        result[current_field] = "\n".join(current_value_lines).strip()

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python add_audio_hooks_tab.py <analysis_name>")
        sys.exit(1)

    analysis_name = sys.argv[1]
    cfg = load_config(analysis_name)
    analysis_dir = BASE / "outputs" / analysis_name

    brand_a_folder = cfg["brand_a"]["folder"]
    brand_b_folder = cfg["brand_b"]["folder"]
    label_a = cfg["brand_a"]["label"]
    label_b = cfg["brand_b"]["label"]
    brands = [(brand_a_folder, label_a), (brand_b_folder, label_b)]

    all_rows = []
    skipped = 0

    for brand_folder, brand_label in brands:
        hook_dir = analysis_dir / "audio_hooks" / brand_folder
        if not hook_dir.exists():
            print(f"[WARN] {brand_folder} hook dir not found")
            continue
        for md_file in sorted(hook_dir.glob("*.md")):
            ad_id = md_file.stem
            content = md_file.read_text(encoding="utf-8")
            if "Hook Analysis Failed" in content:
                skipped += 1
                continue
            fields = extract_hook_fields(content)
            all_rows.append({
                "Brand": brand_label,
                "Ad ID": ad_id,
                "First Spoken Line": fields.get("First Spoken Line", ""),
                "First On-Screen Text": fields.get("First On-Screen Text", ""),
                "Hook Type": fields.get("Hook Type", ""),
                "Hook Summary": fields.get("Hook Summary", ""),
            })

    print(f"Parsed {len(all_rows)} hook analyses (skipped {skipped} failed)")

    xlsx_path = analysis_dir / f"{analysis_name}_analysis.xlsx"
    if not xlsx_path.exists():
        print(f"ERROR: Excel file not found at {xlsx_path}")
        print("Run generate_excel.py first.")
        sys.exit(1)

    wb = load_workbook(str(xlsx_path))

    if "Audio Hooks" in wb.sheetnames:
        del wb["Audio Hooks"]

    ws = wb.create_sheet("Audio Hooks", 5)

    headers = ["Brand", "Ad ID", "First Spoken Line", "First On-Screen Text", "Hook Type", "Hook Summary"]
    HEADER_BG = "FF2C3E50"
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    BG_A = "FFEAF4FB"
    BG_B = "FFE9F7EF"

    for row_idx, row in enumerate(all_rows, 2):
        bg = BG_A if row["Brand"] == label_a else BG_B
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[h])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill("solid", fgColor=bg)

    col_widths = [14, 20, 55, 50, 20, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(str(xlsx_path))
    print(f"Audio Hooks tab added to {xlsx_path}")

    rows_a = [r for r in all_rows if r["Brand"] == label_a]
    rows_b = [r for r in all_rows if r["Brand"] == label_b]
    print(f"\n{label_a} hook types: {dict(Counter(r['Hook Type'] for r in rows_a))}")
    print(f"{label_b} hook types:  {dict(Counter(r['Hook Type'] for r in rows_b))}")

    def safe_print(text):
        sys.stdout.buffer.write((text + "\n").encode("utf-8", errors="replace"))
        sys.stdout.buffer.flush()

    safe_print(f"\nSample {label_a} spoken lines:")
    for r in [x for x in rows_a if x["First Spoken Line"] and x["First Spoken Line"].lower() != "none"][:5]:
        safe_print(f"  [{r['Ad ID']}] {r['First Spoken Line'][:100]}")

    safe_print(f"\nSample {label_b} spoken lines:")
    for r in [x for x in rows_b if x["First Spoken Line"] and x["First Spoken Line"].lower() != "none"][:5]:
        safe_print(f"  [{r['Ad ID']}] {r['First Spoken Line'][:100]}")


if __name__ == "__main__":
    main()
